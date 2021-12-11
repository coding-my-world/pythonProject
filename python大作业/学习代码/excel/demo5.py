# 崔浩然
# 时间:2021/11/17 18:07
# 功能:将字典写入excel表格中
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

data = [
    {
        'name': '小白',
        'tall': 180,
        'age': 23,
        'weight': 74
    },
    {
        'name': '小白',
        'tall': 180,
        'age': 23,
        'weight': 74
    },
    {
        'name': '小白',
        'tall': 180,
        'age': 23,
        'weight': 74
    },
    {
        'name': '小白',
        'tall': 100,
        'age': 23,
        'weight': 74
    }
]

wb = Workbook()
ws = wb.active

title = ['name', 'tall', 'age', 'weight']
ws.append(title)

for person in data:
    ws.append(list(person.values()))

# 计算平均值
for col in range(2, 5):  # range是左闭右开
    char = get_column_letter(col)
    ws[char + '6'] = f"=AVERAGE({char + '2'}:{char + '5'})"

# 修改字体颜色
for col in range(1, 5):
    char = get_column_letter(col)
    ws[char + '1'].font = Font(bold=True, color='000000FF')
wb.save('data.xlsx')
