# 崔浩然
# 时间:2021/11/17 16:32
# 功能：用python操作读取excel表格
from openpyxl import Workbook, load_workbook

wb = load_workbook('excel.xlsx')  # 加载表格

ws = wb.active  # 定位到该表默认的页面

print(ws['A5'].value)
ws['A5'].value = '小兰'  # 修改信息



print(wb.sheetnames)  # 输出工作表名称

wb.create_sheet('qq') #创建工作表

wb.save('excel.xlsx')  # 保存表格信息到文件