# 崔浩然
# 时间:2021/11/17 17:31
# 功能:读取一定范围的excel表格信息以及合并和取消合并单元格
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("new_excel.xlsx")
ws = wb.active

for row in range(1, 5):
    for col in range(1, 6):
        char = get_column_letter(col)  # 获取每列的英文字母
        print(ws[char + str(row)].value)
        
# ws.merged_cells('A1:E1') # 合并单元格
ws.unmerge_cells('A1:E1') # 取消合并单元格
wb.save('new_excel.xlsx')  # 保存工作表
