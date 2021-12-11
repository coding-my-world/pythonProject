# 崔浩然
# 时间:2021/11/23 19:06
# 功能
import random

from openpyxl import Workbook
from openpyxl import load_workbook
from pandas import *


def random_name():
    firstName = "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜戚谢邹喻水云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳鲍史唐费岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅卞齐康伍余元卜顾孟平" \
                "黄和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计成戴宋茅庞熊纪舒屈项祝董粱杜阮席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田胡凌霍万柯卢莫房缪干解应宗丁宣邓郁单杭洪包诸左石崔吉" \
                "龚程邢滑裴陆荣翁荀羊甄家封芮储靳邴松井富乌焦巴弓牧隗山谷车侯伊宁仇祖武符刘景詹束龙叶幸司韶黎乔苍双闻莘劳逄姬冉宰桂牛寿通边燕冀尚农温庄晏瞿茹习鱼容向古戈终居衡步都耿满弘国文东殴沃曾关红游盖益桓公晋楚闫"
    
    # 百家姓中双姓氏
    firstName2 = "万俟司马上官欧阳夏侯诸葛闻人东方赫连皇甫尉迟公羊澹台公冶宗政濮阳淳于单于太叔申屠公孙仲孙轩辕令狐钟离宇文长孙慕容鲜于闾丘司徒司空亓官司寇仉督子颛孙端木巫马公西漆雕乐正壤驷公良拓跋夹谷宰父谷梁段干百里东郭南门呼延羊舌微生梁丘左丘东门西门南宫南宫"
    # 女孩名字
    girl = '秀娟英华慧巧美娜静淑惠珠翠雅芝玉萍红娥玲芬芳燕彩春菊兰凤洁梅琳素云莲真环雪荣爱妹霞香月莺媛艳瑞凡佳嘉琼勤珍贞莉桂娣叶璧璐娅琦晶妍茜秋珊莎锦黛青倩婷姣婉娴瑾颖露瑶怡婵雁蓓纨仪荷丹蓉眉君琴蕊薇菁梦岚苑婕馨瑗琰韵融园艺咏卿聪澜纯毓悦昭冰爽琬茗羽希宁欣飘育滢馥筠柔竹霭凝晓欢霄枫芸菲寒伊亚宜可姬舒影荔枝思丽'
    # 男孩名字
    boy = '伟刚勇毅俊峰强军平保东文辉力明永健世广志义兴良海山仁波宁贵福生龙元全国胜学祥才发武新利清飞彬富顺信子杰涛昌成康星光天达安岩中茂进林有坚和彪博诚先敬震振壮会思群豪心邦承乐绍功松善厚庆磊民友裕河哲江超浩亮政谦亨奇固之轮翰朗伯宏言若鸣朋斌梁栋维启克伦翔旭鹏泽晨辰士以建家致树炎德行时泰盛雄琛钧冠策腾楠榕风航弘'
    # 名
    name = '中笑贝凯歌易仁器义礼智信友上都卡被好无九加电金马钰玉忠孝'
    
    # 10%的机遇生成双数姓氏
    if random.choice(range(100)) > 10:
        firstName_name = firstName[random.choice(range(len(firstName)))]
    else:
        i = random.choice(range(len(firstName2)))
        firstName_name = firstName2[i:i + 2]
    
    sex = random.choice(range(2))           #设置随机性别，概率相等
    name_1 = ""
    # 生成并返回一个名字
    if sex > 0:
        girl_name = girl[random.choice(range(len(girl)))]
        if random.choice(range(2)) > 0:
            name_1 = name[random.choice(range(len(name)))]
        return firstName_name + name_1 + girl_name + "\t女"
    else:
        boy_name = boy[random.choice(range(len(boy)))]
        if random.choice(range(2)) > 0:
            name_1 = name[random.choice(range(len(name)))]
        return firstName_name + name_1 + boy_name + "\t男"


# 生成名单函数
def create_name_list():
    man_list = []
    woman_list = []
    
    for i in range(2200):
        rd_name = ""
        rd_name = random_name()
        if rd_name.split("\t")[1] == '男':
            man_list.append({'name': rd_name.split("\t")[0], 'sex': rd_name.split("\t")[1]})
        elif rd_name.split("\t")[1] == '女':
            woman_list.append({'name': rd_name.split("\t")[0], 'sex': rd_name.split("\t")[1]})
    wb_man = Workbook()
    ws_man = wb_man.active
    wb_woman = Workbook()
    ws_woman = wb_woman.active
    title = ['姓名', '性别', "成绩"]
    ws_man.append(title)
    ws_woman.append(title)
    # 男
    for person in man_list:                               # 分两个循环是为了进行方便成绩的排序，使之后的阳光分班更加方便
        grade = random.randint(500, 700)             # 随机生成成绩，二十八中的学生成绩应该挺好的吧
        person['grade']=grade
    man_list.sort(key=lambda item:item.get("grade"))

    for person in man_list:
        a_name = list(person.values())              # 设置一个列表来存储每个学生的信息
        ws_man.append(list(a_name))
        
    # 女
    for person in woman_list:
        grade = random.randint(500, 700)
        person['grade']=grade
    woman_list.sort(key=lambda item: item.get("grade"))
    for person in woman_list:
        a_name = list(person.values())
        ws_woman.append(list(a_name))
    
    wb_man.save('男生名单.xlsx')
    wb_woman.save('女生名单.xlsx')
    
    
# 分班函数


def fenban():
    # 男
    wb_man = load_workbook("男生名单.xlsx")
    ws_man = wb_man.active                         # 定位到男生名单的默认表格
    man_row=ws_man.max_row-1                       # 获取男生的人数
    per_man=man_row//20                             #每个班的男生数
    last_man=man_row-per_man*20                    #剩余的男生
    
    # 女
    wb_woman = load_workbook("女生名单.xlsx")
    ws_woman = wb_woman.active
    woman_row = ws_woman.max_row-1  # 获取女生的人数
    per_woman = woman_row // 20  # 每个班的女生数
    last_woman = woman_row - per_woman * 20  # 剩余的女生
    
    #分班
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = '1班'
    for i in range(0, 19):
        new_wb.create_sheet(str(i + 2) + '班')
    
    for i in range(0, 20):
        new_ws = new_wb.worksheets[i]
        new_ws.append(['姓名', '性别', '分数'])
        for j in range(0, per_man):
            a_tuple = (ws_man["A" + str(i + 2 + j * 20)].value, ws_man["B" + str(i + 2 + j * 20)].value,ws_man["C" + str(i + 2 + j * 20)].value)  # 用元组将信息存起来
            new_ws.append(list(a_tuple))
        for j in range(0, per_woman):
            a_tuple = (ws_woman["A" + str(i + 2 + j * 20)].value, ws_woman["B" + str(i + 2 + j * 20)].value, ws_woman["C" + str(i + 2 + j * 20)].value)
            new_ws.append(list(a_tuple))
            
    for o in range(0,last_woman+last_man):  #把剩下的学生分到班里
        new_ws = new_wb.worksheets[o%20]
        
        if o>=0 and o<last_woman:
            a_tuple=(ws_woman["A" + str(o + 2 + per_woman * 20)].value, ws_woman["B" + str(o + 2 + per_woman * 20)].value, ws_woman["C" + str(o + 2 + per_woman * 20)].value)
            
            new_ws.append(list(a_tuple))
        elif o>=last_woman and o<last_man+last_woman:
            a_tuple = (ws_man["A" + str(o-last_woman+ 2 + per_man * 20)].value, ws_man["B" + str(o -last_woman+ 2 + per_man * 20)].value, ws_man["C" + str(o -last_woman+ 2 + per_man * 20)].value)
            
            new_ws.append(list(a_tuple))
    new_wb.save('分班名单.xlsx')  # 保存工作表
